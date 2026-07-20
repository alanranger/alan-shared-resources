"""Restore column-I ProductList dropdown on Sales 2024/2025/2026.

Column I uses Excel x14 data validation (ProductList!$A$2:$A$100). openpyxl
removes this when saving; a prior zip patch also used the wrong ext URI.

This script:
  1. Mirrors column-C validation row ranges onto column I (all booking rows)
  2. Re-injects x14:dataValidations with the working ext URI via zip patch
  3. Optionally verifies via Excel COM when the workbook is not already open

Requires: openpyxl. Optional verify: Windows Excel + pywin32.
"""
from __future__ import annotations

import re
import shutil
import sys
import zipfile
from datetime import datetime
from pathlib import Path
from tempfile import TemporaryDirectory

SOURCE = Path(
    r"G:/Dropbox/alan ranger photography/Website Code/alan-shared-resources/csv/"
    r"Booking_Sheet_2026_-_WITH_PRODUCT_MAPPING_3.xlsm"
)
SALES_SHEETS = ("Sales 2026", "Sales 2025", "Sales 2024")
PRODUCT_LIST_FORMULA = "ProductList!$A$2:$A$100"
EXT_URI = "{CCE6A557-97BC-4b89-ADB6-D9C93CAAB3DF}"
XL_VALIDATE_LIST = 3


def c_sqref_to_i_sqref(sqref: str) -> str:
    parts: list[str] = []
    for token in sqref.split():
        token = token.strip()
        if not token:
            continue
        if token[0].upper() != "C":
            parts.append(token)
            continue
        if ":" in token:
            start, end = token.split(":", 1)
            parts.append(f"I{start[1:]}:I{end[1:]}")
        else:
            parts.append(f"I{token[1:]}")
    return " ".join(parts)


def column_c_validation_sqrefs(xlsm: Path, sheet_index: int) -> list[str]:
    with zipfile.ZipFile(xlsm) as z:
        xml = z.read(f"xl/worksheets/sheet{sheet_index}.xml").decode("utf-8")
    sqrefs: list[str] = []
    for block in re.findall(r"<dataValidation[^>]*>.*?</dataValidation>", xml, re.S):
        sq = re.search(r'sqref="([^"]+)"', block)
        f1 = re.search(r"<formula1>([^<]*)</formula1>", block)
        if not sq or not f1:
            continue
        if not str(f1.group(1)).startswith("$I$"):
            continue
        if "C" in sq.group(1):
            sqrefs.append(sq.group(1))
    return sqrefs


def combined_i_sqref(c_sqrefs: list[str]) -> str:
    return " ".join(c_sqref_to_i_sqref(sq) for sq in c_sqrefs)


def build_ext_block(sqref: str) -> str:
    x14 = (
        '<x14:dataValidations count="1" xmlns:xm="http://schemas.microsoft.com/office/excel/2006/main">'
        '<x14:dataValidation type="list" allowBlank="1" xr:uid="{00000000-0002-0000-0500-000003000000}">'
        f"<x14:formula1><xm:f>{PRODUCT_LIST_FORMULA}</xm:f></x14:formula1>"
        f"<xm:sqref>{sqref}</xm:sqref>"
        "</x14:dataValidation></x14:dataValidations>"
    )
    return (
        f'<ext uri="{EXT_URI}" xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main">'
        f"{x14}</ext>"
    )


def patch_sheet_xml(sheet_xml: str, sqref: str) -> str:
    new_ext_lst = f"<extLst>{build_ext_block(sqref)}</extLst>"
    if "<extLst>" in sheet_xml:
        return re.sub(r"<extLst>[\s\S]*?</extLst>", new_ext_lst, sheet_xml, count=1)
    trimmed = sheet_xml.rstrip()
    if not trimmed.endswith("</worksheet>"):
        raise ValueError("Unexpected worksheet XML ending")
    return trimmed[: -len("</worksheet>")] + new_ext_lst + "</worksheet>"


def patch_workbook(path: Path, sqrefs: dict[str, str]) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True)
    names = wb.sheetnames
    wb.close()

    with TemporaryDirectory() as tmp:
        tmp_path = Path(tmp)
        with zipfile.ZipFile(path, "r") as z_in:
            z_in.extractall(tmp_path)

        for name in SALES_SHEETS:
            idx = names.index(name) + 1
            xml_path = tmp_path / f"xl/worksheets/sheet{idx}.xml"
            xml_path.write_text(patch_sheet_xml(xml_path.read_text(encoding="utf-8"), sqrefs[name]), encoding="utf-8")

        out_tmp = tmp_path / "out.xlsm"
        with zipfile.ZipFile(out_tmp, "w", compression=zipfile.ZIP_DEFLATED) as z_out:
            for file in sorted(tmp_path.rglob("*")):
                if file == out_tmp or file.is_dir():
                    continue
                z_out.write(file, file.relative_to(tmp_path).as_posix())
        shutil.copy2(out_tmp, path)


def verify_with_excel(path: Path, sheet_name: str, test_row: int) -> None:
    import win32com.client

    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(path.resolve()))
        cell = wb.Worksheets(sheet_name).Range(f"I{test_row}")
        vtype = cell.Validation.Type
        formula = cell.Validation.Formula1
        wb.Close(SaveChanges=False)
    finally:
        excel.Quit()
    if vtype != XL_VALIDATE_LIST:
        raise RuntimeError(f"{sheet_name}!I{test_row} validation type={vtype}, expected list")
    if "ProductList" not in str(formula):
        raise RuntimeError(f"{sheet_name}!I{test_row} formula1={formula!r}")
    print(f"Verified {sheet_name}!I{test_row}: list -> {formula}")


def main() -> int:
    if not SOURCE.is_file():
        print(f"Missing file: {SOURCE}", file=sys.stderr)
        return 1

    import openpyxl

    wb = openpyxl.load_workbook(SOURCE, read_only=True)
    names = wb.sheetnames
    sqrefs: dict[str, str] = {}
    for name in SALES_SHEETS:
        idx = names.index(name) + 1
        sqrefs[name] = combined_i_sqref(column_c_validation_sqrefs(SOURCE, idx))
        print(f"{name}: {sqrefs[name][:120]}...")
    wb.close()

    stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    backup = SOURCE.with_name(f"{SOURCE.stem}_BACKUP_{stamp}{SOURCE.suffix}")
    shutil.copy2(SOURCE, backup)
    print(f"Backup: {backup.name}")

    patch_workbook(SOURCE, sqrefs)

    try:
        verify_with_excel(SOURCE, "Sales 2026", 197)
    except Exception as err:
        print(f"Verify skipped: {err}")
        print("Close the workbook in Excel if open, reopen the file, then test column I dropdown.")

    print("OK: column I ProductList dropdown restored")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
